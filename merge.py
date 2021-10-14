# import os
import openpyxl
from datetime import datetime
from openpyxl.styles import Alignment, Font

def merge():
    dir = './菜价表'

    # 获取目录下的所有表
    # origin_file_list = 
 
    origin_file_list = ['多多-水果.xlsx', '多多-蔬菜.xlsx', '多多-肉蛋.xlsx',
                        '京东-水果.xlsx', '京东-蔬菜.xlsx', '京东-肉蛋.xlsx',
                        '美团-水果.xlsx', '美团-蔬菜.xlsx', '美团-肉蛋.xlsx']
    
    today = datetime.strftime(datetime.now(), '%Y年%m月%d日')

    wb2 = openpyxl.Workbook()
    for file in origin_file_list:
        file_path = dir + '/' + file
        sheet_name = file.split('/')[-1].split('.')[0]
        print(sheet_name)

        old_wb = openpyxl.load_workbook(file_path)
        old_sheet_name = old_wb.sheetnames[0]
        old_ws = old_wb[old_sheet_name]
        ws2 = wb2.create_sheet(sheet_name)
        for row in old_ws.values:
            ws2.append(row)

        # 单元格格式
        font = Font(size=16, bold=True)
        align = Alignment(horizontal='center',vertical='center',wrap_text=True)

        ws2.merge_cells('A1:C1')
        ws_style = ws2['A1']
        ws_style.font = font
        ws_style.alignment = align
        ws2.column_dimensions['A'].width=32
        ws2.row_dimensions[1].height=25

    del wb2['Sheet']
    wb2.save('./菜价表/每日菜价/'+ today + '线上平台菜价.xlsx')
    print("完成！")
