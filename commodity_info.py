import json
import re
# import pandas as pd
from jsonpath import jsonpath
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

def info():
    while True:
        
        ch = int(input("输入序号（京喜拼拼：1，美团优选：2，多多买菜：3，输入0退出。）："))
        if ch == 0:
            break
        
        sku = []
        price1 = []
        sale = []
        while True:
            file = input("输入文件名(不带后缀名，输入q继续。)：")

            # 退出循环
            if file =='q':
                break

            # 文件名拼接
            read_file = './json/' + file + ".json"
            save_file = './prices_excel/' + file + ".xlsx"

            # 读取json文件
            with open(read_file, 'r' ,encoding='utf-8') as f:
                data = json.load(f)

            # 提取需要的内容
            # data = data['data']s
            # data = data['pageList']时
            if ch == 1:
                sku_name = jsonpath(data, '$..pageList[*].skuName')
                price = jsonpath(data, '$..pageList[*].price')
                sale_price = jsonpath(data, '$..pageList[*].jdPrice')
            elif ch == 2:
                sku_name = jsonpath(data, '$..itemList[*].skuTitle[text]')
                price = jsonpath(data, '$..itemList[*].sellPrice[text]')
                sale_price = jsonpath(data, '$..itemList[*].newSellPrice[text]')
            elif ch == 3:
                sku_name = jsonpath(data, '$..goods_list[*].goods_name')
                price = jsonpath(data, '$..goods_list[*].market_price')
                sale_price = jsonpath(data, '$..goods_list[*].price')
            else:
                break
            sku += sku_name 
            price1 += price
            sale += sale_price 

            
                
        # data_raw = pd.DataFrame(columns=data.keys())
        # data_raw = data_raw.append(data,ignore_index=True)
        # print(data_raw)

        # print(sku_name, price, sale_price)

        wb = Workbook()
        # 工作表0
        ws = wb.active
        ws.title = "01" # sheet名
        
        # ws = wb.create_sheet('date') # 创建工作表

        # 填写数据
        font = Font(size=16, bold=True)
        align = Alignment(horizontal='center',vertical='center',wrap_text=True)

        # ws['A1'].alignment = align
        if ch == 1:
            ws['A1'] = '京喜拼拼'
        elif ch == 2:
            ws['A1'] = '美团优选'
        elif ch == 3:
            ws['A1'] = '多多买菜'
        ws.merge_cells('A1:C1')
        ws_style = ws['A1']
        ws_style.font = font
        ws_style.alignment = align
        ws['A2'] = '品名'
        ws['B2'] = '正常价格'
        ws['C2'] = '优惠价格'

        # 设置列宽、行高
        ws.column_dimensions['A'].width=32
        ws.row_dimensions[1].height=25

        val1 = 3
        val2 = 3
        val3 = 3
    
        pattern = re.compile(r'[-+]?[0-9]*\.?[0-9]+')

        for n in sku:
            ws.cell(row=val1, column=1, value=n)
            val1 += 1
        for p in price1:
            if isinstance(p, str):
                p1 = pattern.findall(p)
                ws.cell(row=val2, column=2, value=float(p1[0]))
            elif isinstance(p, int): # 多多优选
                p1 = float(p) / 100
                ws.cell(row=val2, column=2, value=p1)
            val2 += 1
        for s in sale:
            if isinstance(s, str):
                s1 = pattern.findall(s)
                ws.cell(row=val3, column=3, value=float(s1[0]))
            elif isinstance(s, int): # 多多优选
                s1 = float(s) / 100
                ws.cell(row=val3, column=3, value=s1)

            val3 += 1

        # 保存
        wb.save(save_file)
        print("OK!")